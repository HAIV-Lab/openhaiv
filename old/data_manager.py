import torch
import logging
import numpy as np

from torch.utils.data import Dataset
from torchvision import transforms
from openpyxl import load_workbook
from collections import defaultdict
from .data import DATASETS


class DataManager(object):
    """manage the increment datasets

    Args:
        dataset_name (str): name of the increment dataset
        dataset_path (str): path of the increment dataset
        shuffle (bool): shuffle the class order by using np.random.permutation
        seed (int): random seed of numpy
        init_cls (int): num of base classes
        increment (int): num of classes increased
        use_path (bool): whether DummyDataset.images is a list of image paths(str)
        crop_transform (CustomMultiCropping): instance of CustomMultiCropping, return of the function util.get_transform()
        secondary_transform (list): list of torchvision.transforms, return of the function util.get_transform()
        attr_path (str): path of the attribute.xlsx
    
    Examples:
        >>> from ncdia.data import get_transform, DataManager
        >>> cil_path = "./CIL_Dataset"
        >>> datamanager = DataManager("remote", cil_path, shuffle=False, seed=0, init_cls=10, increment=2)
        
        >>> crop_transform, secondary_transform = get_transform(cfg.dataloader)
        >>> datamanager = DataManager("remote", cil_path, shuffle=False, seed=0, init_cls=10, increment=2, use_path=True, 
                                      crop_transform=crop_transform, secondary_transform=secondary_transform, attr_path=os.path.join(cil_path, "Attribute0421.xlsx"))
    """
    def __init__(self, dataset_name: str, dataset_path: str, shuffle: bool, seed: int, init_cls: int, increment: int, use_path=False,
                 crop_transform=None, secondary_transform=None, attr_path=""):
        self.dataset_name = dataset_name
        self.dataset_path = dataset_path
        self.nbase = init_cls
        self._setup_data(shuffle, seed)
        assert init_cls <= len(self._class_order), "No enough classes."
        self._increments = [init_cls]

        while sum(self._increments) + increment < len(self._class_order):
            self._increments.append(increment)
        offset = len(self._class_order) - sum(self._increments)
        if offset > 0:
            self._increments.append(offset)
        
        self.use_path = use_path
        self.attr_path = attr_path
        self.crop_transform = crop_transform
        self.secondary_transform = secondary_transform

    @property
    def nb_tasks(self):
        return len(self._increments)

    def get_task_size(self, task):
        return self._increments[task]
    
    def get_accumulate_tasksize(self,task):
        return sum(self._increments[:task+1])
    
    def get_total_classnum(self):
        return len(self._class_order)

    def _get_data(self, source: str):
        if source == "train":
            x, y = self._train_data, self._train_targets
        elif source == "test":
            x, y = self._test_data, self._test_targets
        elif source == "ncd":
            x, y = self._ncd_data, self._ncd_targets
        else:
            raise ValueError("Unknown data source {}.".format(source))
        return x, y
    
    def _get_trsf(self, mode: str):
        if mode == "train":
            trsf = transforms.Compose([*self._train_trsf, *self._common_trsf])
        elif mode == "test":
            trsf = transforms.Compose([*self._test_trsf, *self._common_trsf])
        elif mode == "ncd":
            trsf = transforms.Compose([*self._ncd_trsf, *self._common_trsf])
        else:
            raise ValueError("Unknown mode {}.".format(mode))
        return trsf
    
    def get_dataset(self, indices, source: str, mode: str, appendent=None, ret_data=False, m_rate=None):
        """get the increment datasets

        Args:
            indices: 1D numpy array or list of int, select and return the data of certain classes
            source (str): select from train data, test data and ncd data
            mode (str): select transform from train mode, test mode and ncd mode
            appendent: None or tuple of additional dataset with (appendent_data(numpy.ndarray), appendent_targets(numpy.ndarray))
            ret_data (bool): whether to return the concat data and targets (numpy.ndarray)
            m_rate: None or float in range (0, 1), select number of classes with (1 - m_rate) * len([low_range, high_range))
        
        Returns:
            if ret_data: return data, targets, dataset (DummyDataset)
            else: return dataset (DummyDataset)
        """
        x, y = self._get_data(source)
        trsf = self._get_trsf(mode)

        data, targets = [], []
        for idx in indices:
            if m_rate is None:
                class_data, class_targets = self._select(
                    x, y, low_range=idx, high_range=idx + 1
                )
            else:
                class_data, class_targets = self._select_rmm(
                    x, y, low_range=idx, high_range=idx + 1, m_rate=m_rate
                )
            data.append(class_data)
            targets.append(class_targets)

        if appendent is not None and len(appendent) != 0:
            appendent_data, appendent_targets = appendent
            data.append(appendent_data)
            targets.append(appendent_targets)

        data, targets = np.concatenate(data), np.concatenate(targets)

        if len(self.attr_path) > 0:
            dataset = DummyAttrDataset(data, targets, trsf, self.use_path, self.crop_transform, self.secondary_transform, self.attr_path)
        else:
            dataset = DummyDataset(data, targets, trsf, self.use_path)
        
        if ret_data:
            return data, targets, dataset
        else:
            return dataset

        
    def get_finetune_dataset(self, known_classes: int, total_classes: int, source: str, mode: str, appendent: tuple, val_samples_per_class: int=0):
        """get the validation datasets

        Args:
            known_classes (int): number of known classes
            total_classes (int): number of total classes
            source (str): select from train data, test data and ncd data
            mode (str): select transform from train mode, test mode and ncd mode
            appendent (tuple): (appendent_data, appendent_targets)
            val_samples_per_class (int): how many data per class will be sampled from the unknown dataset and added to the validation dataset

        Returns:
            dataset (DummyDataset): validation dataset with selected data and transforms
        """
        x, y = self._get_data(source)
        trsf = self._get_trsf(mode)
        
        val_data = []
        val_targets = []
        appendent_data, appendent_targets = appendent

        for idx in range(0, known_classes):
            append_data, append_targets = self._select(appendent_data, appendent_targets,
                                                       low_range=idx, high_range=idx+1)
            if len(append_data) > 0:
                val_data.append(append_data)
                val_targets.append(append_targets)
        
        # if type == "ratio":
        #     new_num_tot = old_num_tot * (total_classes-known_classes) // known_classes
        # elif type == "same":
        #     new_num_tot = old_num_tot
        # else:
        #     assert 0, "not implemented yet"
        
        for idx in range(known_classes, total_classes):
            class_data, class_targets = self._select(x, y, low_range=idx, high_range=idx+1)
            val_indx = np.random.choice(len(class_data), val_samples_per_class, replace=False)
            val_data.append(class_data[val_indx])
            val_targets.append(class_targets[val_indx])
        val_data, val_targets = np.concatenate(val_data), np.concatenate(val_targets)

        if len(self.attr_path) > 0:
            dataset = DummyAttrDataset(val_data, val_targets, trsf, self.use_path, self.crop_transform, self.secondary_transform, self.attr_path)
        else:
            dataset = DummyDataset(val_data, val_targets, trsf, self.use_path)
        return dataset

    def get_dataset_with_split(self, indices, source: str, mode: str, appendent=None, val_samples_per_class: int=0):
        """get both of the train dataset and the validation dataset

        Args: 
            indices: 1D numpy array or list of int, select and return the data of certain classes
            source (str): select from train data, test data and ncd data
            mode (str): select transform from train mode, test mode and ncd mode
            appendent: None or tuple of additional dataset with (appendent_data(numpy.ndarray), appendent_targets(numpy.ndarray))
            val_samples_per_class (int): how many data per class will be sampled from the whole dataset and added to the validation dataset

        Returns: 
            train_dataset (DummyDataset): train dataset
            val_dataset (DummyDataset): validation dataset
        """
        x, y = self._get_data(source)
        trsf = self._get_trsf(mode)

        train_data, train_targets = [], []
        val_data, val_targets = [], []
        for idx in indices:
            class_data, class_targets = self._select(
                x, y, low_range=idx, high_range=idx + 1
            )
            val_indx = np.random.choice(
                len(class_data), val_samples_per_class, replace=False
            )
            train_indx = list(set(np.arange(len(class_data))) - set(val_indx))
            val_data.append(class_data[val_indx])
            val_targets.append(class_targets[val_indx])
            train_data.append(class_data[train_indx])
            train_targets.append(class_targets[train_indx])

        if appendent is not None:
            appendent_data, appendent_targets = appendent
            for idx in range(0, int(np.max(appendent_targets)) + 1):
                append_data, append_targets = self._select(
                    appendent_data, appendent_targets, low_range=idx, high_range=idx + 1
                )
                val_indx = np.random.choice(
                    len(append_data), val_samples_per_class, replace=False
                )
                train_indx = list(set(np.arange(len(append_data))) - set(val_indx))
                val_data.append(append_data[val_indx])
                val_targets.append(append_targets[val_indx])
                train_data.append(append_data[train_indx])
                train_targets.append(append_targets[train_indx])

        train_data, train_targets = np.concatenate(train_data), np.concatenate(train_targets)
        val_data, val_targets = np.concatenate(val_data), np.concatenate(val_targets)

        if len(self.attr_path) > 0:
            train_dataset = DummyAttrDataset(train_data, train_targets, trsf, self.use_path, self.crop_transform, self.secondary_transform, self.attr_path)
            val_dataset = DummyAttrDataset(val_data, val_targets, trsf, self.use_path, self.crop_transform, self.secondary_transform, self.attr_path)
        else:
            train_dataset = DummyDataset(train_data, train_targets, trsf, self.use_path)
            val_dataset = DummyDataset(val_data, val_targets, trsf, self.use_path)
        
        return train_dataset, val_dataset

    def _setup_data(self, shuffle, seed):
        idata = _get_idata(self.dataset_name, self.dataset_path)
        idata.download_data()

        # Data
        self._train_data, self._train_targets = idata.train_data, idata.train_targets
        self._test_data, self._test_targets = idata.test_data, idata.test_targets
        self._ncd_data, self._ncd_targets = idata.ncd_data, idata.ncd_targets
        self.use_path = idata.use_path

        # Transforms
        self._train_trsf = idata.train_trsf
        self._test_trsf = idata.test_trsf
        self._ncd_trsf = idata.ncd_trsf
        self._common_trsf = idata.common_trsf

        # Order
        order = [i for i in range(len(np.unique(self._train_targets)))]
        if shuffle:
            np.random.seed(seed)
            order = np.random.permutation(len(order)).tolist()
        else:
            order = idata.class_order
        self._class_order = order
        logging.info(self._class_order)

        # Map indices
        self._train_targets = _map_new_class_index(self._train_targets, self._class_order)
        self._test_targets = _map_new_class_index(self._test_targets, self._class_order)
        self._ncd_targets = _map_new_class_index(self._ncd_targets, self._class_order)

    def _select(self, x, y, low_range, high_range):
        idxes = np.where(np.logical_and(y >= low_range, y < high_range))[0]
        
        if isinstance(x,np.ndarray):
            x_return = x[idxes]
        else:
            x_return = []
            for id in idxes:
                x_return.append(x[id])
        return x_return, y[idxes]

    def _select_rmm(self, x, y, low_range, high_range, m_rate):
        assert m_rate is not None
        if m_rate != 0:
            idxes = np.where(np.logical_and(y >= low_range, y < high_range))[0]
            selected_idxes = np.random.randint(
                0, len(idxes), size=int((1 - m_rate) * len(idxes))
            )
            new_idxes = idxes[selected_idxes]
            new_idxes = np.sort(new_idxes)
        else:
            new_idxes = np.where(np.logical_and(y >= low_range, y < high_range))[0]
        return x[new_idxes], y[new_idxes]

    def getlen(self, index):
        y = self._train_targets
        return np.sum(np.where(y == index))


class DummyDataset(Dataset):
    """subclass of torch.utils.data.Dataset

    Args:
        images (list): list of images in the dataset
        labels (list): list of labels in the dataset
        trsf (transforms.Compose): the transforms applied to the images
        use_path (bool): whether DummyDataset.images is a list of image paths(str)

    Examples:
        >>> ds = DummyDataset(images, labels, trsf, use_path)
        >>> print(len(ds))
        >>> print(ds[0])
        {
            'data': ..., 
            'label': ..., 
            'imgpath': ...
        }
    """
    def __init__(self, images: list, labels: list, trsf: transforms.Compose, use_path=False):
        assert len(images) == len(labels), "Data size error!"
        self.images = images
        self.labels = labels
        self.trsf = trsf
        self.use_path = use_path

    def __len__(self):
        return len(self.images)

    def __getitem__(self, idx):
        if self.use_path:
            image = pil_loader(self.images[idx])
        else:
            image = Image.fromarray(self.images[idx])
        data = self.trsf(image)
        label = self.labels[idx]
        
        return {
            'data': data, 
            'label': label, 
            'imgpath': self.images[idx] if self.use_path else ""
        }


class DummyAttrDataset(DummyDataset):
    """DummyDataset with crop transform, secondary transform and attributes

    Args:
        images (list): list of images in the dataset
        labels (list): list of labels in the dataset
        trsf (transforms.Compose): the transforms applied to the images
        use_path (bool): whether DummyDataset.images is a list of image paths(str)
        crop_transform (CustomMultiCropping): instance of CustomMultiCropping, return of the function util.get_transform()
        secondary_transform (list): list of torchvision.transforms, return of the function util.get_transform()
        attr_path (str): path of the attribute.xlsx

    Examples:
        >>> from ncdia.data import get_transform
        >>> crop_transform, secondary_transform = get_transform(cfg.dataloader)
        >>> ds = DummyAttrDataset(images, labels, trsf, use_path, crop_transform, secondary_transform, attr_path)
        >>> print(len(ds))
        >>> print(ds[0])
        {
            'data': ..., 
            'label': ...,
            'attribute': ..., 
            'imgpath': ...
        }
    """
    def __init__(self, images: list, labels: list, trsf: transforms.Compose, use_path=False, crop_transform=None, secondary_transform=None, attr_path=""):
        super().__init__(images, labels, trsf, use_path)
        self.attr_path = attr_path
        self.crop_transform = crop_transform
        self.secondary_transform = secondary_transform
        if isinstance(secondary_transform, list):
            assert (len(secondary_transform) == self.crop_transform.N_large + self.crop_transform.N_small)
        
        wb = load_workbook(self.attr_path)
        sheets = wb.worksheets
        sheet1 = sheets[0]
        max_row_num = sheet1.max_row
        self.attr_dict = defaultdict(list)
        for i in range(2, max_row_num + 1):
            row_list2 = []
            count = 0
            for row in sheet1[i]:
                if count == 0:
                    tag = row.value
                    count += 1
                    continue
                if row.value is None:
                    continue
                row_list2.append(row.value)
                count += 1
            self.attr_dict[tag] = row_list2
    
    def __getitem__(self, idx):
        label = self.labels[idx]
        if self.use_path:
            image = pil_loader(self.images[idx])
        else:
            image = Image.fromarray(self.images[idx])
        
        if self.crop_transform is not None and self.secondary_transform is not None:
            classify_image = [self.trsf(image)]
            multi_crop, multi_crop_params = self.crop_transform(image)
            assert (len(multi_crop) == self.crop_transform.N_large + self.crop_transform.N_small)
            if isinstance(self.secondary_transform, list):
                multi_crop = [tf(x) for tf, x in zip(self.secondary_transform, multi_crop)]
            else:
                multi_crop = [self.secondary_transform(x) for x in multi_crop]
            data = classify_image + multi_crop
        else:
            data = self.trsf(image)

        attrs = self.attr_dict[label]
        attrs = torch.tensor(list(map(float, attrs)), dtype=torch.float)

        return {
            'data': data, 
            'label': label,
            'attribute': attrs, 
            'imgpath': self.images[idx] if self.use_path else ""
        }


def _map_new_class_index(y, order):
    return None if y is None else np.array(list(map(lambda x: order.index(x), y)))


def _get_idata(dataset_name, path=""):
    name = dataset_name.lower()
    target = DATASETS.modules()[name]
    return target(path)
