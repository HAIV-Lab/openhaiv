import os
import random
import torch
from torchvision import transforms

from PIL import Image
from openpyxl import load_workbook
from ncdia.utils import DATASETS
from ncdia.dataloader.tools import pil_loader
from .utils import BaseDataset
from ncdia.dataloader.augmentations.constrained_cropping import CustomMultiCropping

from collections import defaultdict
attr_path = "/data/datasets/remote/0718_dy/Attribute0718.xlsx"

def get_transform():
    normalize = transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                         std=[0.229, 0.224, 0.225])
    num_crops = [2, 4]
    size_crops = [224, 96]
    min_scale_crops =  [0.14, 0.05]
    max_scale_crops = [1, 0.14]
    constrained_cropping = False
    crop_transform = CustomMultiCropping(size_large=size_crops[0],
                                         scale_large=(min_scale_crops[0], max_scale_crops[0]),
                                         size_small=size_crops[1],
                                         scale_small=(min_scale_crops[1], max_scale_crops[1]),
                                         N_large=num_crops[0], N_small=num_crops[1],
                                         condition_small_crops_on_key=constrained_cropping)
    secondary_transform = transforms.Compose([
        transforms.RandomHorizontalFlip(),    
        transforms.RandomApply([
        transforms.ColorJitter(0.4, 0.4, 0.4, 0.1)
        ], p=0.8),
        transforms.RandomGrayscale(p=0.2),
        transforms.ToTensor(),
        normalize]
        )
    return crop_transform, secondary_transform


@DATASETS.register
class Remote(BaseDataset):
    """Remote dataset

    Args:
        root (str): root folder of the dataset
        split (str): split of the dataset. Should be one of 'train', 'test'.
        subset_labels (list | int): subset of labels
        subset_file (str): file containing the selected images
        transform (list | str): transform to apply on the dataset.
            If str, it should be one of 'train', 'test' for predefined transforms.

    Attributes:
        images (list): list of image paths
        labels (list): list of labels
        num_classes (int): number of classes
        transform (torchvision.transforms.Compose): transform to apply on the dataset
    
    """
    train_transform = transforms.Compose([
        transforms.Resize(256, interpolation=transforms.InterpolationMode.BILINEAR),
        transforms.RandomResizedCrop(224),
        transforms.RandomHorizontalFlip(),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
    ])

    test_transform = transforms.Compose([
        transforms.Resize(256, interpolation=transforms.InterpolationMode.BILINEAR),
        transforms.CenterCrop(224),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])

    def __init__(
            self,
            root: str,
            split: str = 'train',
            subset_labels: list = None,
            subset_file: str = None,
            transform: list | str = None,
            **kwargs,
    ) -> None:
        super().__init__()
        self.root = root
        self.split = split

        if split == 'train':
            self.images, self.labels = self._load_data(os.path.join(root, 'train'))
        elif split == 'test':
            self.images, self.labels = self._load_data(os.path.join(root, 'test'))
        else:
            raise ValueError(f"Unknown split: {split}")
        
        if subset_labels is not None:
            self.images, self.labels = self._select_from_label(self.images, self.labels, subset_labels)
        if subset_file is not None:
            self.images, self.labels = self._select_from_file(self.images, self.labels, subset_file)

        if isinstance(transform, str):
            if transform == 'train':
                self.transform = self.train_transform
            elif transform == 'test':
                self.transform = self.test_transform
            else:
                raise ValueError(f"Unknown transform: {transform}")
        else:
            self.transform = transform
        self.crop_transform, self.secondary_transform = get_transform()
        self.multi_train = False
        self.attr_dict = defaultdict(list)
        self._load_att()

    def _load_att(self):
        wb = load_workbook(attr_path)
        sheets = wb.worksheets
        sheet1 = sheets[0]
        max_row_num = sheet1.max_row
        for i in range(2, max_row_num + 1):
            row_list2 = []
            count = 0
            for row in sheet1[i]:
                if count == 0:
                    tag = row.value
                    count += 1
                    continue
                if row.value is None:
                    continue
                row_list2.append(row.value)
                count += 1
            self.attr_dict[tag] = row_list2

    def _load_data(self, root: str):
        """Load data from root folder

        Args:
            root (str): root folder of the dataset

        Returns:
            list: list of image paths
            list: list of labels
        """
        imgpaths, labels = [], []
        for num, dir_name in enumerate(os.listdir(root)):
            img_list = os.listdir(os.path.join(root, dir_name))
            for k in range(len(img_list)):
                imgpath = os.path.join(root, dir_name, img_list[k])
                imgpaths.append(os.path.abspath(imgpath))
                # if int(dir_name) >= 33:
                if int(num) >= 33:
                    labels.append(num)
                else:
                    labels.append(num)
        return imgpaths, labels
    
    def _select_from_label(self, images: list, labels: list, subset_labels: list | int):
        """Select images from a subset of labels

        Args:
            images (list): list of image paths
            labels (list): list of labels
            subset_labels (list | int): list of subset labels

        Returns:
            list: list of image paths
            list: list of labels
        """
        if isinstance(subset_labels, int):
            subset_labels = [i for i in range(subset_labels)]
        selected_images, selected_labels = [], []
        for img, label in zip(images, labels):
            if label in subset_labels:
                selected_images.append(img)
                selected_labels.append(label)
        return selected_images, selected_labels

    def _select_from_file(self, images: list, labels: list, file: str):
        """Select images from a subset of labels

        Args:
            images (list): list of image paths
            labels (list): list of labels
            file (str): file containing the selected images

        Returns:
            list: list of image paths
            list: list of labels
        """
        selected_images, selected_labels = [], []
        with open(file, 'r') as f:
            for line in f:
                img = os.path.abspath(line.strip())
                if img in images:
                    selected_images.append(img)
                    selected_labels.append(labels[images.index(img)])
        return selected_images, selected_labels
    
    def sample_test_data(self, ratio: int | float):
        """Sample a subset of the test data for each class according to the given ratio.

        Args:
            ratio (int | float): ratio of the sampled data.
                If int, it represents the number of samples to be selected.
                If float, it represents the ratio of the samples to be selected.
        
        Returns:
            Sampled data and corresponding targets.
        """
        if self.split != 'test':
            raise ValueError("Sampling is only applicable in test mode")

        class_indices = {}
        for idx, label in enumerate(self.targets):
            if label in class_indices:
                class_indices[label].append(idx)
            else:
                class_indices[label] = [idx]

        sampled_images, sampled_labels = [], []
        for label, indices in class_indices.items():
            if isinstance(ratio, int) and ratio > 0:
                sample_size = min(len(indices), ratio)
            elif isinstance(ratio, float) and 0 < ratio < 1:
                sample_size = max(1, int(len(indices) * ratio))
            else:
                raise ValueError("Invalid ratio value")
            
            sampled_indices = random.sample(indices, sample_size)
            for i in sampled_indices:
                sampled_images.append(self.images[i])
                sampled_labels.append(self.labels[i])

        self.images, self.labels = sampled_images, sampled_labels

    def __len__(self) -> int:
        return len(self.images)

    def __getitem__(self, index: int) -> dict:
        imgpath, label = self.images[index], self.labels[index]
        img = pil_loader(imgpath)
        if self.multi_train:
            image = Image.open(imgpath).convert('RGB')
            classify_image = [self.transform(image)]
            # print(self.crop_transform(image))
            multi_crop = self.crop_transform(image)
            assert (len(multi_crop) == self.crop_transform.N_large + self.crop_transform.N_small)
            if isinstance(self.secondary_transform, list):
                multi_crop = [tf(x) for tf, x in zip(self.secondary_transform, multi_crop)]
            else:
                multi_crop = [self.secondary_transform(x) for x in multi_crop]
            total_img = classify_image + multi_crop

        if self.transform is not None:
            total_img = self.transform(img)
        attrs = self.attr_dict[label]
        attrs = torch.tensor(list(map(float, attrs)), dtype=torch.float)

        return {
            'data': total_img,
            'label': label,
            'attribute': attrs,
            'imgpath': imgpath,
        }
